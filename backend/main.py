import os
import uuid
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from io import BytesIO

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, Form, HTTPException, Query, Request
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

import crud
from auth import create_access_token, get_current_user, require_roles
from database import engine, get_db
from models import Base, User
from schemas import (
    BlockCreate,
    BlockListResponse,
    BlockResponse,
    BlockUpdate,
    DashboardMTDStats,
    DashboardResponse,
    DashboardTodayStats,
    HaulingTransactionCreate,
    HaulingTransactionEnvelope,
    HaulingTransactionListResponse,
    HaulingTransactionUpdate,
    ItemCreate,
    ItemListResponse,
    ItemResponse,
    ItemStatsResponse,
    ItemUpdate,
    LoginRequest,
    TokenResponse,
    UserCreate,
    UserResponse,
    VendorCreate,
    VendorListResponse,
    VendorResponse,
    VendorUpdate,
)

load_dotenv()

# Buat semua tabel
Base.metadata.create_all(bind=engine)


def ensure_hauling_schema() -> None:
    """Tambahkan kolom/objek hauling baru jika database sudah terlanjur ada."""
    ddl_statements = [
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS role VARCHAR(20) NOT NULL DEFAULT 'operator'",
        "ALTER TABLE hauling_transactions ADD COLUMN IF NOT EXISTS transaction_date DATE",
        "ALTER TABLE hauling_transactions ADD COLUMN IF NOT EXISTS notes TEXT",
        "ALTER TABLE hauling_transactions ADD COLUMN IF NOT EXISTS created_by INTEGER",
        "ALTER TABLE hauling_transactions ADD COLUMN IF NOT EXISTS updated_by INTEGER",
        "ALTER TABLE hauling_transactions ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP WITH TIME ZONE",
        "ALTER TABLE hauling_transactions ADD COLUMN IF NOT EXISTS deleted_by INTEGER",
        "ALTER TABLE hauling_transactions ALTER COLUMN ticket_no TYPE VARCHAR(100)",
        "ALTER TABLE hauling_transactions ALTER COLUMN vehicle_plate TYPE VARCHAR(20)",
        "ALTER TABLE items ADD COLUMN IF NOT EXISTS price FLOAT DEFAULT 0",
    ]

    with engine.begin() as connection:
        for statement in ddl_statements:
            try:
                connection.execute(text(statement))
            except Exception:
                pass


ensure_hauling_schema()

EXPORT_RATE_LIMIT_WINDOW_SECONDS = 60
EXPORT_RATE_LIMIT_MAX_REQUESTS = 10
export_request_history: dict[int, deque[datetime]] = defaultdict(deque)

app = FastAPI(
    title="PalmChain API v2",
    description="REST API untuk Palm Oil Supply Chain Monitoring System (PalmChain)",
    version="1.0.0",
    docs_url="/docs",
    openapi_url="/openapi.json",
)

# ==================== CORS (FIXED) ====================
allowed_origins = os.getenv("ALLOWED_ORIGINS", "http://localhost:5173")
origins_list = [origin.strip() for origin in allowed_origins.split(",")]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _error_payload(status_code: int, message: str, code: str | None = None, errors: dict | None = None):
    payload = {"success": False, "message": message}
    if code:
        payload["code"] = code
    if errors:
        payload["errors"] = errors
    return JSONResponse(status_code=status_code, content=payload)


@app.exception_handler(HTTPException)
def http_exception_handler(request: Request, exc: HTTPException):
    code_map = {
        400: "BAD_REQUEST",
        401: "UNAUTHORIZED",
        403: "FORBIDDEN",
        404: "NOT_FOUND",
        429: "TOO_MANY_REQUESTS",
    }
    code = code_map.get(exc.status_code, "BAD_REQUEST" if exc.status_code < 500 else "INTERNAL_ERROR")
    return _error_payload(exc.status_code, str(exc.detail), code=code)


@app.exception_handler(RequestValidationError)
def validation_exception_handler(request: Request, exc: RequestValidationError):
    field_errors: dict[str, list[str]] = {}
    for error in exc.errors():
        location = error.get("loc", [])
        field_name = location[-1] if location else "detail"
        field_errors.setdefault(str(field_name), []).append(error.get("msg", "Validasi gagal"))
    return _error_payload(422, "Validasi gagal", errors=field_errors)


@app.exception_handler(Exception)
def unhandled_exception_handler(request: Request, exc: Exception):
    return _error_payload(500, "Terjadi kesalahan pada server", code="INTERNAL_ERROR")


# ==================== HEALTH CHECK ====================

@app.get("/health")
def health_check():
    return {"status": "healthy", "version": "0.4.0"}


# ==================== AUTH ENDPOINTS (PUBLIC) ====================

@app.post("/auth/register", response_model=UserResponse, status_code=201)
def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """
    Registrasi user baru.

    Requirements:
    - **email**: Format valid (contoh: user@itk.ac.id)
    - **name**: Minimal 2 karakter, maksimal 100 karakter
    - **password**: Minimal 8 karakter dengan kombinasi:
      - Huruf (A-Z, a-z)
      - Angka (0-9)
      - Special character (!@#$%^&*)

    Contoh password valid: `Password123!`
    """
    try:
        user = crud.create_user(db=db, user_data=user_data)
        if not user:
            raise HTTPException(
                status_code=400,
                detail="Email sudah terdaftar. Gunakan email lain atau login dengan akun yang sudah ada."
            )
        return user
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/auth/login", response_model=TokenResponse)
def login(login_data: LoginRequest, db: Session = Depends(get_db)):
    """
    Login dengan JSON body dan dapatkan JWT token.

    - **email**: Email pengguna yang terdaftar
    - **password**: Password pengguna

    **Response:**
    - **access_token**: JWT token untuk otorisasi (valid 60 menit)
    - **token_type**: Tipe token (selalu 'bearer')
    - **user**: Informasi user yang login

    **Penggunaan:**
    Gunakan token di header setiap request:
    ```
    Authorization: Bearer <access_token>
    ```
    """
    user = crud.authenticate_user(db=db, email=login_data.email, password=login_data.password)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Email atau password salah. Periksa kembali kredensial Anda."
        )

    token = create_access_token(data={"sub": user.id})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": user,
    }


@app.post("/auth/token")
def login_for_access_token(
    username: str = Form(..., description="Email user"),
    password: str = Form(..., description="Password user"),
    db: Session = Depends(get_db),
):
    """
    OAuth2 compatible token endpoint.

    Endpoint ini menerima form data (username=email, password=password).
    Digunakan oleh Swagger UI Authorization dan OAuth2 clients.
    """
    # Username di OAuth2 context berarti email
    user = crud.authenticate_user(db=db, email=username, password=password)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Email atau password salah",
            headers={"WWW-Authenticate": "Bearer"},
        )

    token = create_access_token(data={"sub": user.id})
    return {
        "access_token": token,
        "token_type": "bearer",
    }


@app.get("/auth/me", response_model=UserResponse)
def get_me(current_user: User = Depends(get_current_user)):
    """Ambil profil user yang sedang login."""
    return current_user


# ==================== VENDOR ENDPOINTS (PROTECTED) ====================

@app.post("/api/vendors", response_model=VendorResponse, status_code=201)
def create_vendor(
    vendor: VendorCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Buat vendor baru. **Membutuhkan autentikasi.**"""
    db_vendor = crud.create_vendor(db=db, vendor_data=vendor)
    if not db_vendor:
        raise HTTPException(status_code=400, detail="Vendor code sudah terdaftar")
    return db_vendor


@app.get("/api/vendors", response_model=VendorListResponse)
def list_vendors(
    skip: int | None = Query(None, ge=0),
    limit: int | None = Query(None, ge=1, le=1000),
    search: str = Query(None),
    status: bool = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil daftar vendors dengan pagination atau mode dropdown. **Membutuhkan autentikasi.**"""
    if skip is None and limit is None:
        vendors = crud.get_vendors(db=db, skip=0, limit=10000, search=search, status=True if status is None else status)
        return {
            "success": True,
            "data": [{"id": vendor.id, "name": vendor.name, "code": vendor.code} for vendor in vendors["vendors"]],
            "total": vendors["total"],
            "vendors": vendors["vendors"],
        }

    return crud.get_vendors(db=db, skip=skip or 0, limit=limit or 20, search=search, status=status)


@app.get("/api/vendors/{vendor_id}", response_model=VendorResponse)
def get_vendor(
    vendor_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil satu vendor berdasarkan ID. **Membutuhkan autentikasi.**"""
    try:
        vendor_uuid = uuid.UUID(vendor_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid vendor ID format")

    vendor = crud.get_vendor(db=db, vendor_id=vendor_uuid)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor tidak ditemukan")
    return vendor


@app.put("/api/vendors/{vendor_id}", response_model=VendorResponse)
def update_vendor(
    vendor_id: str,
    vendor: VendorUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update vendor. **Membutuhkan autentikasi.**"""
    try:
        vendor_uuid = uuid.UUID(vendor_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid vendor ID format")

    updated = crud.update_vendor(db=db, vendor_id=vendor_uuid, vendor_data=vendor)
    if not updated:
        raise HTTPException(status_code=404, detail="Vendor tidak ditemukan")
    return updated


@app.delete("/api/vendors/{vendor_id}", status_code=204)
def delete_vendor(
    vendor_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Hapus vendor. **Membutuhkan autentikasi.**"""
    try:
        vendor_uuid = uuid.UUID(vendor_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid vendor ID format")

    success = crud.delete_vendor(db=db, vendor_id=vendor_uuid)
    if not success:
        raise HTTPException(status_code=404, detail="Vendor tidak ditemukan")


# ==================== BLOCK ENDPOINTS (PROTECTED) ====================

@app.post("/api/blocks", response_model=BlockResponse, status_code=201)
def create_block(
    block: BlockCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Buat block/afdeling baru. **Membutuhkan autentikasi.**"""
    db_block = crud.create_block(db=db, block_data=block)
    if not db_block:
        raise HTTPException(status_code=400, detail="Block code sudah terdaftar")
    return db_block


@app.get("/api/blocks", response_model=BlockListResponse)
def list_blocks(
    skip: int | None = Query(None, ge=0),
    limit: int | None = Query(None, ge=1, le=1000),
    search: str = Query(None),
    status: bool = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil daftar blocks dengan pagination atau mode dropdown. **Membutuhkan autentikasi.**"""
    if skip is None and limit is None:
        blocks = crud.get_blocks(db=db, skip=0, limit=10000, search=search, status=True if status is None else status)
        return {
            "success": True,
            "data": [
                {"id": block.id, "name": block.division or block.block_code, "code": block.block_code, "area_ha": block.hectarage}
                for block in blocks["blocks"]
            ],
            "total": blocks["total"],
            "blocks": blocks["blocks"],
        }

    return crud.get_blocks(db=db, skip=skip or 0, limit=limit or 20, search=search, status=status)


@app.get("/api/blocks/{block_id}", response_model=BlockResponse)
def get_block(
    block_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil satu block berdasarkan ID. **Membutuhkan autentikasi.**"""
    try:
        block_uuid = uuid.UUID(block_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid block ID format")

    block = crud.get_block(db=db, block_id=block_uuid)
    if not block:
        raise HTTPException(status_code=404, detail="Block tidak ditemukan")
    return block


@app.put("/api/blocks/{block_id}", response_model=BlockResponse)
def update_block(
    block_id: str,
    block: BlockUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update block. **Membutuhkan autentikasi.**"""
    try:
        block_uuid = uuid.UUID(block_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid block ID format")

    updated = crud.update_block(db=db, block_id=block_uuid, block_data=block)
    if not updated:
        raise HTTPException(status_code=404, detail="Block tidak ditemukan")
    return updated


@app.delete("/api/blocks/{block_id}", status_code=204)
def delete_block(
    block_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Hapus block. **Membutuhkan autentikasi.**"""
    try:
        block_uuid = uuid.UUID(block_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid block ID format")

    success = crud.delete_block(db=db, block_id=block_uuid)
    if not success:
        raise HTTPException(status_code=404, detail="Block tidak ditemukan")


# ==================== HAULING TRANSACTION ENDPOINTS (PROTECTED) ====================

@app.post("/api/hauling-transactions", response_model=HaulingTransactionEnvelope, status_code=201)
def create_hauling_transaction(
    hauling: HaulingTransactionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "supervisor", "operator")),
):
    """Buat hauling transaction baru. **Membutuhkan autentikasi.**"""
    db_hauling, error_code = crud.create_hauling_transaction(db=db, hauling_data=hauling, created_by=current_user)
    if not db_hauling:
        if error_code == "duplicate_ticket":
            raise HTTPException(status_code=400, detail="Ticket number sudah terdaftar")
        if error_code == "invalid_vendor":
            raise HTTPException(status_code=400, detail="Vendor tidak ditemukan atau tidak aktif")
        if error_code == "invalid_block":
            raise HTTPException(status_code=400, detail="Block tidak ditemukan atau tidak aktif")
        if error_code == "invalid_weight":
            raise HTTPException(status_code=400, detail="Weight Out tidak boleh melebihi Weight In")
        raise HTTPException(status_code=400, detail="Gagal menyimpan transaksi hauling")
    return {"success": True, "data": db_hauling}


@app.get("/api/hauling-transactions", response_model=HaulingTransactionListResponse)
def list_hauling_transactions(
    ticket_no: str = Query(None, max_length=100),
    vendor_id: str = Query(None),
    block_id: str = Query(None),
    date_from: date = Query(None),
    date_to: date = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=5, le=100),
    sort_by: str = Query("transaction_date"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil daftar hauling transactions dengan pagination dan filter. **Membutuhkan autentikasi.**"""
    vendor_uuid = None
    block_uuid = None

    if vendor_id:
        try:
            vendor_uuid = uuid.UUID(vendor_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid vendor ID format")

    if block_id:
        try:
            block_uuid = uuid.UUID(block_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid block ID format")

    return crud.get_hauling_transactions(
        db=db,
        vendor_id=vendor_uuid,
        block_id=block_uuid,
        ticket_no=ticket_no,
        date_from=date_from,
        date_to=date_to,
        page=page,
        per_page=per_page,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )


@app.get("/api/hauling-transactions/{hauling_id}", response_model=HaulingTransactionEnvelope)
def get_hauling_transaction(
    hauling_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil satu hauling transaction berdasarkan ID. **Membutuhkan autentikasi.**"""
    try:
        hauling_uuid = uuid.UUID(hauling_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid hauling ID format")

    hauling = crud.get_hauling_transaction(db=db, hauling_id=hauling_uuid)
    if not hauling:
        raise HTTPException(status_code=404, detail="Hauling transaction tidak ditemukan")
    return {"success": True, "data": hauling}


@app.put("/api/hauling-transactions/{hauling_id}", response_model=HaulingTransactionEnvelope)
def update_hauling_transaction(
    hauling_id: str,
    hauling: HaulingTransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "supervisor", "operator")),
):
    """Update hauling transaction. **Membutuhkan autentikasi.**"""
    try:
        hauling_uuid = uuid.UUID(hauling_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid hauling ID format")

    updated, error_code = crud.update_hauling_transaction(db=db, hauling_id=hauling_uuid, hauling_data=hauling, updated_by=current_user)
    if not updated:
        if error_code == "duplicate_ticket":
            raise HTTPException(status_code=400, detail="Ticket number sudah terdaftar")
        if error_code == "invalid_vendor":
            raise HTTPException(status_code=400, detail="Vendor tidak ditemukan atau tidak aktif")
        if error_code == "invalid_block":
            raise HTTPException(status_code=400, detail="Block tidak ditemukan atau tidak aktif")
        if error_code == "invalid_weight":
            raise HTTPException(status_code=400, detail="Weight Out tidak boleh melebihi Weight In")
        raise HTTPException(status_code=404, detail="Hauling transaction tidak ditemukan")
    return {"success": True, "data": updated}


@app.delete("/api/hauling-transactions/{hauling_id}", status_code=200)
def delete_hauling_transaction(
    hauling_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    """Hapus hauling transaction. **Membutuhkan autentikasi.**"""
    try:
        hauling_uuid = uuid.UUID(hauling_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid hauling ID format")

    success = crud.delete_hauling_transaction(db=db, hauling_id=hauling_uuid, deleted_by=current_user)
    if not success:
        raise HTTPException(status_code=404, detail="Hauling transaction tidak ditemukan")

    return {"success": True, "message": "Transaksi berhasil dihapus"}


def _enforce_export_rate_limit(user_id: int) -> None:
    now = datetime.now()
    history = export_request_history[user_id]
    window_start = now - timedelta(seconds=EXPORT_RATE_LIMIT_WINDOW_SECONDS)

    while history and history[0] < window_start:
        history.popleft()

    if len(history) >= EXPORT_RATE_LIMIT_MAX_REQUESTS:
        raise HTTPException(status_code=429, detail="Terlalu banyak request export. Coba lagi sebentar lagi.")

    history.append(now)


@app.get("/api/hauling-transactions/export")
def export_hauling_transactions(
    format: str = Query(..., pattern="^(xlsx|pdf)$"),
    ticket_no: str = Query(None, max_length=100),
    vendor_id: str = Query(None),
    block_id: str = Query(None),
    date_from: date = Query(None),
    date_to: date = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "supervisor", "manager")),
):
    """Export hauling transactions sesuai filter aktif."""
    _enforce_export_rate_limit(current_user.id)

    vendor_uuid = None
    block_uuid = None

    if vendor_id:
        try:
            vendor_uuid = uuid.UUID(vendor_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid vendor ID format")

    if block_id:
        try:
            block_uuid = uuid.UUID(block_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid block ID format")

    result = crud.get_hauling_transactions(
        db=db,
        page=1,
        per_page=10000,
        ticket_no=ticket_no,
        vendor_id=vendor_uuid,
        block_id=block_uuid,
        date_from=date_from,
        date_to=date_to,
        sort_by="transaction_date",
        sort_dir="desc",
    )

    if result["meta"]["total"] > 10000:
        raise HTTPException(status_code=400, detail="Data terlalu banyak, gunakan filter untuk mempersempit hasil")

    rows = result["data"]
    export_date = datetime.now().strftime("%Y-%m-%d")

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hauling Transactions"
        headers = ["No.", "Ticket No", "Vehicle Plate", "Vendor", "Block Code", "Weight In", "Weight Out", "Net Weight", "Tanggal", "Catatan"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")

        for index, row in enumerate(rows, start=1):
            sheet.append([
                index,
                row["ticket_no"],
                row["vehicle_plate"],
                None if row["vendor"] is None else row["vendor"]["name"],
                None if row["block"] is None else row["block"]["code"],
                row["weight_in"],
                row["weight_out"],
                row["net_weight"],
                row["transaction_date"],
                row["notes"],
            ])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="hauling-transactions-{export_date}.xlsx"'},
        )

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        raise HTTPException(status_code=500, detail="PDF export dependency belum tersedia")

    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph("Laporan Hauling TBS", styles["Title"]), Spacer(1, 8)]
    table_data = [["No.", "Ticket No", "Vehicle Plate", "Vendor", "Block Code", "Weight In", "Weight Out", "Net Weight", "Tanggal", "Catatan"]]
    for index, row in enumerate(rows, start=1):
        table_data.append([
            index,
            row["ticket_no"],
            row["vehicle_plate"],
            "" if row["vendor"] is None else row["vendor"]["name"],
            "" if row["block"] is None else row["block"]["code"],
            f"{row['weight_in']:.3f}",
            f"{row['weight_out']:.3f}",
            f"{row['net_weight']:.3f}",
            row["transaction_date"].strftime("%d/%m/%Y") if row["transaction_date"] else "",
            row["notes"] or "",
        ])
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(table)
    document.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="hauling-transactions-{export_date}.pdf"'})


# ==================== ITEM ENDPOINTS (PROTECTED) ====================

@app.post("/api/items", response_model=ItemResponse, status_code=201)
def create_item(
    item: ItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Buat item baru. **Membutuhkan autentikasi.**"""
    db_item = crud.create_item(db=db, item_data=item)
    if not db_item:
        raise HTTPException(status_code=400, detail="Item code sudah terdaftar")
    return db_item


@app.get("/api/items", response_model=ItemListResponse)
def list_items(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=1000),
    search: str = Query(None),
    category: str = Query(None),
    status: bool = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Ambil daftar items dengan pagination. **Membutuhkan autentikasi.**

    Query parameters:
    - **category**: Filter berdasarkan kategori (e.g., 'electronics', 'hardware')
    - **search**: Cari berdasarkan code atau name
    - **status**: Filter by active status
    """
    return crud.get_items(db=db, skip=skip, limit=limit, search=search, category=category, status=status)


@app.get("/api/items/stats", response_model=ItemStatsResponse)
def get_item_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil statistik item berdasarkan kategori. **Membutuhkan autentikasi.**"""
    return crud.get_item_stats(db=db)


@app.get("/api/items/{item_id}", response_model=ItemResponse)
def get_item(
    item_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ambil satu item berdasarkan ID. **Membutuhkan autentikasi.**"""
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid item ID format")

    item = crud.get_item(db=db, item_id=item_uuid)
    if not item:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")
    return item


@app.put("/api/items/{item_id}", response_model=ItemResponse)
def update_item(
    item_id: str,
    item: ItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update item. **Membutuhkan autentikasi.**"""
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid item ID format")

    updated = crud.update_item(db=db, item_id=item_uuid, item_data=item)
    if not updated:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")
    return updated


@app.delete("/api/items/{item_id}", status_code=204)
def delete_item(
    item_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Hapus item. **Membutuhkan autentikasi.**"""
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid item ID format")

    success = crud.delete_item(db=db, item_id=item_uuid)
    if not success:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")


# ==================== DASHBOARD ENDPOINTS (PROTECTED) ====================

@app.get("/api/dashboard", response_model=DashboardResponse)
def get_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Dapatkan dashboard summary dengan statistik hari ini dan Month-To-Date.

    **Response:**
    - **today**: Statistik transaksi hari ini (total_transactions, total_tonage, avg_tonage)
    - **mtd**: Statistik bulan ini (total_transactions, total_tonage, target_tonage, achievement_percentage)
    - **last_updated**: Waktu terakhir data diperbarui

    **Membutuhkan autentikasi.**
    """
    from datetime import datetime

    today_stats = crud.get_hauling_stats_today(db=db)
    mtd_stats = crud.get_hauling_stats_mtd(db=db, target_tonage=500.0)

    return {
        "today": DashboardTodayStats(**today_stats),
        "mtd": DashboardMTDStats(**mtd_stats),
        "last_updated": datetime.now()
    }


# ==================== TEAM INFO ====================
@app.get("/team")
def team_info():
    return {
        "team": "cloud-team-a-awit",
        "members": [
            {"name": "Adam Ibnu Ramadhan", "nim": "10231003", "role": "Lead Backend"},
            {"name": "Alfian Fadillah Putra", "nim": "10231009", "role": "Lead Frontend"},
            {"name": "Varrel Kaleb Ropard Pasaribu", "nim": "10231089", "role": "Lead Container"},
            {"name": "Adhyasta Firdaus", "nim": "10231005", "role": "Lead CI/CD & Deploy"},
            {"name": "Adonia Azarya Tamalonggehe", "nim": "10231007", "role": "Lead QA & Docs"},
        ]
    }
